#!/usr/bin/env python3
"""
Portfolio tracker — generates Portfolio_Tracker.xlsx with near-real-time prices.

Reads positions.yaml (own portfolio) + father_portfolio.yaml, fetches live
prices via yfinance (NSE + US), and writes a formatted multi-tab Excel
workbook. When yfinance is unreachable (offline / proxy), prices fall back
to prices_override.json (if fresh) or the last values stored in the YAMLs,
and each row is labeled live / override / cached so you always know what
you're looking at.

Usage:
    pip install yfinance pyyaml openpyxl
    python portfolio_tracker.py                 # writes Portfolio_Tracker.xlsx
    python portfolio_tracker.py --no-fetch      # skip yfinance, use cached
    python portfolio_tracker.py --out my.xlsx

Refresh workflow:
    Re-run the script any time; the file is regenerated in place with a
    timestamp on the Dashboard. Schedule it (cron / launchd) for hands-free
    daily updates.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import yaml
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Missing dependency: {e.name}. Install with: pip install pyyaml openpyxl yfinance")

# yfinance is optional — tracker degrades to cached prices without it
try:
    import yfinance as yf
    HAVE_YF = True
except ImportError:
    HAVE_YF = False

HERE = Path(__file__).parent

# Broker symbol → Yahoo Finance symbol
YF_MAP = {
    # Zerodha / INDmoney ETFs & REITs
    "NIFTYBEES": "NIFTYBEES.NS", "BANKBEES": "BANKBEES.NS",
    "JUNIORBEES": "JUNIORBEES.NS", "SILVERBEES": "SILVERBEES.NS",
    "MID150BEES": "MID150BEES.NS", "GOLDBEES": "GOLDBEES.NS",
    "ITBEES": "ITBEES.NS", "PSUBNKBEES": "PSUBNKBEES.NS",
    "MINDSPACE-RR": "MINDSPACE.NS", "EMBASSY": "EMBASSY.NS",
    "BIRET": "BIRET.NS", "NXST": "NXST.NS",
    # Indian stocks (note: Gulf Oil Lubricants is GULFOILLUB, not GULFPETRO)
    "HINDALCO": "HINDALCO.NS", "GULFPETRO": "GULFOILLUB.NS",
    "SONACOMS": "SONACOMS.NS", "MOTHERSON": "MOTHERSON.NS", "NETWEB": "NETWEB.NS",
    # Father's holdings
    "NESTLEIND": "NESTLEIND.NS", "COLPAL": "COLPAL.NS", "GILLETTE": "GILLETTE.NS",
    # US tickers are NOT listed here — they equal their own Yahoo symbol,
    # and Prices.get() falls back to the raw symbol when it's absent above.
}

# ── styling helpers ─────────────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1F2937")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SEC_FILL = PatternFill("solid", fgColor="E5E7EB")
SEC_FONT = Font(bold=True, size=11)
GREEN = Font(color="15803D")
RED = Font(color="B91C1C")
DIM = Font(color="6B7280", size=9)
THIN = Border(bottom=Side(style="thin", color="D1D5DB"))

ACTION_FILL = {
    "HOLD": PatternFill("solid", fgColor="DCFCE7"),
    "BUY": PatternFill("solid", fgColor="DBEAFE"),
    "SELL_PLANNED": PatternFill("solid", fgColor="FEE2E2"),
    "HOLD_FOR_TLH": PatternFill("solid", fgColor="FEF3C7"),
    "MOSTLY_SOLD": PatternFill("solid", fgColor="F3F4F6"),
    "SOLD": PatternFill("solid", fgColor="F3F4F6"),
}
INPUT_FILL = PatternFill("solid", fgColor="FEF9C3")   # yellow = type here
CALC_FILL = PatternFill("solid", fgColor="F0F9FF")    # blue = computed


def header_row(ws, row, cols):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    return row + 1


def section(ws, row, title, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill, cell.font = SEC_FILL, SEC_FONT
    return row + 1


def pnl_font(cell, val):
    if val is None:
        return
    cell.font = GREEN if val >= 0 else RED


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── price engine ────────────────────────────────────────────────────────

class Prices:
    def __init__(self, fetch: bool):
        self.overrides = {}
        ov_path = HERE / "prices_override.json"
        if ov_path.exists():
            raw = json.loads(ov_path.read_text())
            self.overrides = {k: v for k, v in raw.items() if not k.startswith("_")}
            self.override_time = raw.get("_updated", "?")
        self.fetch = fetch and HAVE_YF
        self.cache = {}

    def get(self, symbol: str, cached_ltp: float | None) -> tuple[float | None, float | None, str]:
        """Return (price, day_pct, source).

        Only "live" and "override" results are cached by symbol — those are
        genuinely the same market price no matter which account/holding
        asks. A "cached" fallback is NOT cached globally: it's derived from
        the caller's own cached_ltp (e.g. that position's cost basis), which
        differs per account even for the same symbol (same ticker held in
        two different accounts at two different avg prices). Caching it by
        symbol alone would leak one account's cost basis into another's.
        """
        if symbol in self.cache:
            return self.cache[symbol]
        result = None
        if self.fetch:
            # NSE symbols need YF_MAP's .NS suffix; US tickers equal themselves
            yf_symbol = YF_MAP.get(symbol, symbol)
            try:
                h = yf.Ticker(yf_symbol).history(period="5d", interval="1d")
                if h is not None and len(h) >= 1:
                    last = float(h["Close"].iloc[-1])
                    prev = float(h["Close"].iloc[-2]) if len(h) > 1 else last
                    result = (last, (last / prev - 1) * 100, "live")
            except Exception:
                result = None
        if result is None and symbol in self.overrides:
            o = self.overrides[symbol]
            result = (o["price"], o.get("day_pct"), "override")
        if result is not None:
            self.cache[symbol] = result
            return result
        return (cached_ltp, None, "cached")


# ── sheet builders ──────────────────────────────────────────────────────

def build_holdings_sheet(ws, title, groups, prices: Prices):
    """groups: list of (section_title, rows) where each row is a dict with
    symbol, name, qty, avg (optional), ltp (cached fallback)."""
    ws.sheet_view.showGridLines = False
    cols = ["Symbol", "Name", "Qty", "Avg Cost", "LTP", "Day %",
            "Invested", "Value", "P&L", "P&L %", "Price Src"]
    row = 1
    grand_value = grand_invested = 0.0
    for sec_title, items in groups:
        row = section(ws, row, sec_title, len(cols))
        row = header_row(ws, row, cols)
        sec_value = sec_invested = 0.0
        for it in items:
            sym = it["symbol"]
            qty = it.get("qty")
            avg = it.get("avg")
            price, day_pct, src = prices.get(sym, it.get("ltp"))
            value = qty * price if (qty and price) else it.get("value")
            invested = (qty * avg) if (qty and avg) else it.get("invested")
            pnl = (value - invested) if (value is not None and invested is not None) else None
            pnl_pct = (pnl / invested * 100) if (pnl is not None and invested) else None

            vals = [sym, it.get("name", ""), qty, avg, price, day_pct,
                    invested, value, pnl, pnl_pct, src]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = THIN
                if ci in (4, 5, 7, 8, 9):
                    cell.number_format = "#,##0.00"
                if ci in (6, 10):
                    cell.number_format = "0.00%" if False else "+0.00;-0.00"
                if ci == 9:
                    pnl_font(cell, pnl)
                if ci == 10 and pnl_pct is not None:
                    pnl_font(cell, pnl_pct)
                if ci == 11:
                    cell.font = DIM
            if value:
                sec_value += value
            if invested:
                sec_invested += invested
            row += 1
        # section subtotal
        ws.cell(row=row, column=2, value="Subtotal").font = Font(bold=True)
        c = ws.cell(row=row, column=7, value=sec_invested or None)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        c = ws.cell(row=row, column=8, value=sec_value or None)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        if sec_invested:
            c = ws.cell(row=row, column=9, value=sec_value - sec_invested)
            c.number_format = "#,##0"
            pnl_font(c, sec_value - sec_invested)
            c.font = Font(bold=True, color=c.font.color)
        row += 2
        grand_value += sec_value
        grand_invested += sec_invested
    autosize(ws, [14, 34, 8, 11, 11, 8, 13, 13, 12, 9, 9])
    ws.freeze_panes = "A2"
    return grand_value, grand_invested


def build_mf_sheet(ws, funds, totals):
    ws.sheet_view.showGridLines = False
    row = section(ws, 1, "Mutual Funds (NAV as of last INDmoney sync — refresh via app export)", 8)
    row = header_row(ws, row, ["Fund", "Units", "NAV", "Invested", "Current", "P&L", "P&L %", "Via"])
    for f in funds:
        pnl_pct = f["pnl"] / f["invested"] * 100 if f.get("invested") else None
        vals = [f["fund"], f["units"], f["nav"], f["invested"], f["current"], f["pnl"], pnl_pct, f.get("via", "")]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN
            if ci in (2, 3):
                cell.number_format = "#,##0.00"
            if ci in (4, 5, 6):
                cell.number_format = "#,##0"
            if ci == 6:
                pnl_font(cell, f["pnl"])
            if ci == 7 and pnl_pct is not None:
                cell.number_format = "+0.0;-0.0"
                pnl_font(cell, pnl_pct)
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, key in ((4, "total_invested"), (5, "total_current"), (6, "total_pnl")):
        c = ws.cell(row=row, column=ci, value=totals.get(key))
        c.number_format = "#,##0"
        c.font = Font(bold=True)
    autosize(ws, [42, 11, 10, 11, 11, 10, 8, 13])
    ws.freeze_panes = "A3"


def build_us_sheet(ws, accounts, fx, prices: Prices):
    """accounts: list of (account_label, us_dict) — each rendered as its own
    labeled block (ETFs + Stocks sub-sections) with a subtotal, so multiple
    US brokers/accounts can share one tab without conflating their totals."""
    ws.sheet_view.showGridLines = False
    cols = ["Symbol", "Name", "Units", "Price $", "Value ₹", "Invested ₹", "P&L ₹", "P&L %", "Price Src"]
    row = 1
    grand_value = grand_invested = 0.0
    for account_label, us in accounts:
        acct_value = acct_invested = 0.0
        for sec_title, items in ((f"{account_label} — ETFs", us.get("etfs", [])),
                                  (f"{account_label} — Stocks", us.get("stocks", []))):
            if not items:
                continue
            row = section(ws, row, sec_title, len(cols))
            row = header_row(ws, row, cols)
            for it in items:
                units = it["units"]
                # New schema: qty + avg (USD cost basis). Old schema fallback: value_inr/pnl.
                if "avg" in it:
                    avg_cost = it["avg"]
                    price, _, src = prices.get(it["symbol"], avg_cost)
                    invested = units * avg_cost * fx
                else:
                    cached_px = (it["value_inr"] / units / fx) if units else None
                    price, _, src = prices.get(it["symbol"], cached_px)
                    invested = it["value_inr"] - it["pnl"]
                value_inr = units * price * fx if price else invested
                pnl = value_inr - invested
                pnl_pct = pnl / invested * 100 if invested else None
                vals = [it["symbol"], it.get("name", ""), units, price, value_inr, invested, pnl, pnl_pct, src]
                for ci, v in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    cell.border = THIN
                    if ci == 3:
                        cell.number_format = "0.000"
                    if ci == 4:
                        cell.number_format = "#,##0.00"
                    if ci in (5, 6, 7):
                        cell.number_format = "#,##0"
                    if ci == 7:
                        pnl_font(cell, pnl)
                    if ci == 8 and pnl_pct is not None:
                        cell.number_format = "+0.0;-0.0"
                        pnl_font(cell, pnl_pct)
                    if ci == 9:
                        cell.font = DIM
                acct_value += value_inr
                acct_invested += invested
                row += 1
            row += 1
        # account subtotal
        ws.cell(row=row, column=2, value=f"{account_label} subtotal").font = Font(bold=True)
        c = ws.cell(row=row, column=6, value=acct_invested)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        c = ws.cell(row=row, column=5, value=acct_value)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        c = ws.cell(row=row, column=7, value=acct_value - acct_invested)
        c.number_format = "#,##0"
        pnl_font(c, acct_value - acct_invested)
        c.font = Font(bold=True, color=c.font.color)
        row += 2
        grand_value += acct_value
        grand_invested += acct_invested
    autosize(ws, [9, 34, 9, 10, 12, 12, 11, 8, 9])
    ws.freeze_panes = "A3"
    return grand_value, grand_invested
    return grand_value, grand_invested


def build_swing_sheet(ws, cfg, prices: Prices):
    ws.sheet_view.showGridLines = False
    row = section(ws, 1, "Open Swing Positions", 9)
    open_pos = [p for p in cfg.get("positions", []) if p.get("status") == "open"]
    if open_pos:
        row = header_row(ws, row, ["Ticker", "Qty", "Entry", "LTP", "Stop", "Next Target", "P&L %", "Risk ₹", "Src"])
        for p in open_pos:
            # Look up by the plain ticker (matches YF_MAP and prices_override.json
            # convention used everywhere else) — NOT yf_symbol, which would bypass
            # both and only work for live yfinance fetch, silently missing overrides.
            price, _, src = prices.get(p["ticker"], p.get("entry_avg"))
            pnl_pct = (price / p["entry_avg"] - 1) * 100 if price else None
            next_t = next((t["price"] for t in p.get("targets", []) if price and t["price"] > price), None)
            risk = max(0, (p["entry_avg"] - p["stop"])) * p["qty"]
            vals = [p["ticker"], p["qty"], p["entry_avg"], price, p["stop"], next_t, pnl_pct, risk, src]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = THIN
                if ci == 7 and pnl_pct is not None:
                    cell.number_format = "+0.0;-0.0"
                    pnl_font(cell, pnl_pct)
            row += 1
    else:
        ws.cell(row=row, column=1, value="No open positions — flat.").font = DIM
        row += 1
    row += 1

    row = section(ws, row, "Planned Positions (staged, not filled)", 9)
    planned_pos = [p for p in cfg.get("positions", []) if p.get("status") == "planned"]
    if planned_pos:
        row = header_row(ws, row, ["Ticker", "Plan Qty", "Pilot Entry", "LTP", "Stop", "Pivot", "Dist to Buy Zone", "Risk ₹ (full)", "Src"])
        for p in planned_pos:
            price, _, src = prices.get(p["ticker"], p.get("entry_avg"))
            zone = p.get("buy_range")
            if zone and price:
                lo, hi = zone
                dist = "IN ZONE" if lo <= price <= hi else (f"{(lo/price-1)*100:+.1f}% to zone" if price < lo else f"{(price/hi-1)*100:+.1f}% past zone")
            else:
                dist = ""
            risk = max(0, (p["entry_avg"] - p["stop"])) * p.get("qty", 0)
            vals = [p["ticker"], p.get("qty"), p.get("entry_avg"), price, p["stop"], p.get("pivot"), dist, risk, src]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = THIN
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
            row += 1
        # planned position notes below the table
        for p in planned_pos:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            ws.cell(row=row, column=1, value=f"↳ {p['ticker']}: {p.get('add_plan', '')}").font = DIM
            row += 1
    else:
        ws.cell(row=row, column=1, value="No planned positions.").font = DIM
        row += 1
    row += 1

    row = section(ws, row, "Watchlist / Planned Entries", 9)
    row = header_row(ws, row, ["Ticker", "Trigger", "Stop", "Note", "", "", "", "", ""])
    for w in cfg.get("watchlist", []):
        ws.cell(row=row, column=1, value=w.get("ticker")).border = THIN
        ws.cell(row=row, column=2, value=str(w.get("trigger", ""))).border = THIN
        ws.cell(row=row, column=3, value=str(w.get("stop", ""))).border = THIN
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
        ws.cell(row=row, column=4, value=w.get("note", "")).font = DIM
        row += 1
    row += 1

    row = section(ws, row, "Closed Trade Journal", 9)
    row = header_row(ws, row, ["Ticker", "Result", "Avg Entry", "Avg Exit", "P&L", "P&L %", "Hold Days", "", ""])
    for c in cfg.get("closed", []):
        s = c.get("summary", {})
        pnl = s.get("pnl_inr", s.get("pnl_usd"))
        cur = "₹" if "pnl_inr" in s else "$"
        vals = [c["ticker"], c.get("result", ""), s.get("avg_entry"), s.get("avg_exit"),
                f"{cur}{pnl:,.0f}" if pnl is not None else "", s.get("pnl_pct"), s.get("hold_days")]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN
            if ci == 6 and isinstance(v, (int, float)):
                cell.number_format = "+0.0;-0.0"
                pnl_font(cell, v)
        row += 1
    autosize(ws, [12, 12, 11, 11, 11, 9, 10, 10, 8])


def build_father_sheet(ws, fp, prices: Prices):
    ws.sheet_view.showGridLines = False
    row = section(ws, 1, f"Father's Portfolio — as of {fp.get('as_of', '?')} (LTP refreshed on generation)", 10)
    row = header_row(ws, row, ["Symbol", "Name", "Qty", "Avg Cost", "LTP", "Value",
                               "Invested", "P&L", "P&L %", "Action"])
    tot_v = tot_i = 0.0
    for h in fp.get("holdings", []):
        qty, avg = h.get("qty"), h.get("avg")
        price, _, src = prices.get(h["symbol"], None)
        value = qty * price if (qty and price) else None
        invested = h.get("invested") or ((qty * avg) if (qty and avg) else None)
        pnl = value - invested if (value is not None and invested is not None) else None
        pnl_pct = pnl / invested * 100 if (pnl is not None and invested) else None
        vals = [h["symbol"], h.get("name", ""), qty if qty is not None else "TBD",
                avg, price, value, invested, pnl, pnl_pct, h.get("action", "")]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN
            if ci in (4, 5):
                cell.number_format = "#,##0.00"
            if ci in (6, 7, 8):
                cell.number_format = "#,##0"
            if ci == 8:
                pnl_font(cell, pnl)
            if ci == 9 and pnl_pct is not None:
                cell.number_format = "+0.0;-0.0"
                pnl_font(cell, pnl_pct)
            if ci == 10:
                cell.fill = ACTION_FILL.get(str(v), PatternFill())
                cell.font = Font(bold=True, size=9)
        if value:
            tot_v += value
        if invested:
            tot_i += invested
        # note line
        row += 1
        if h.get("note"):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
            ws.cell(row=row, column=2, value="↳ " + h["note"]).font = DIM
            row += 1
    ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
    for ci, v in ((6, tot_v), (7, tot_i), (8, tot_v - tot_i if tot_i else None)):
        c = ws.cell(row=row, column=ci, value=v or None)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        if ci == 8 and v is not None:
            pnl_font(c, v)
    row += 2
    row = section(ws, row, "Planned Trades (redeployment of sale proceeds)", 10)
    planned = fp.get("planned_trades") or []
    if planned:
        row = header_row(ws, row, ["Ticker", "Entry Plan", "Stop", "Size", "Note", "", "", "", "", ""])
        for p in planned:
            ws.cell(row=row, column=1, value=p.get("ticker")).border = THIN
            ws.cell(row=row, column=2, value=str(p.get("entry", ""))).border = THIN
            ws.cell(row=row, column=3, value=str(p.get("stop", ""))).border = THIN
            ws.cell(row=row, column=4, value=str(p.get("size", ""))).border = THIN
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
            ws.cell(row=row, column=5, value=p.get("note", "")).font = DIM
            row += 1
    else:
        ws.cell(row=row, column=1, value="None yet — ~₹50L expected free after GILLETTE + COLPAL exits. Plan before redeploying.").font = DIM
    autosize(ws, [12, 28, 8, 11, 11, 13, 13, 12, 9, 14])


def build_calc_sheet(ws, sleeve_inr):
    """Interactive position-sizing + R:R calculator driven by Excel formulas.
    Yellow cells = inputs; blue cells = computed. Mirrors rr_calc.py logic."""
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="POSITION SIZING & R:R CALCULATOR").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Yellow = type your numbers. Blue = auto-computed. Mirrors rr_calc.py.").font = DIM

    def label(r, c, text, bold=False):
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = Font(bold=bold, size=10)
        return cell

    def inp(r, c, val, fmt="#,##0.00"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = INPUT_FILL
        cell.number_format = fmt
        cell.border = THIN
        return cell

    def calc(r, c, formula, fmt="#,##0.00"):
        cell = ws.cell(row=r, column=c, value=formula)
        cell.fill = CALC_FILL
        cell.number_format = fmt
        cell.border = THIN
        return cell

    # ── inputs ──
    r = 4
    section(ws, r, "Trade Inputs", 4)
    label(r + 1, 1, "Entry price");            inp(r + 1, 2, 100.00)
    label(r + 2, 1, "Stop loss");              inp(r + 2, 2, 92.00)
    label(r + 3, 1, "Account / sleeve ₹");     inp(r + 3, 2, sleeve_inr, "#,##0")
    label(r + 4, 1, "Risk budget %");          inp(r + 4, 2, 2.0, "0.0")
    # B5=entry B6=stop B7=sleeve B8=risk%

    # ── sizing ──
    r = 10
    section(ws, r, "Position Sizing (from risk budget)", 4)
    label(r + 1, 1, "Risk per share");         calc(r + 1, 2, "=B5-B6")
    label(r + 2, 1, "Risk % of entry");        calc(r + 2, 2, "=(B5-B6)/B5", "0.00%")
    label(r + 3, 1, "Max ₹ at risk");          calc(r + 3, 2, "=B7*B8/100", "#,##0")
    label(r + 4, 1, "Max shares");             calc(r + 4, 2, "=IF(B11>0,FLOOR(B13/B11,1),0)", "#,##0")
    label(r + 5, 1, "Capital required");       calc(r + 5, 2, "=B14*B5", "#,##0")
    label(r + 6, 1, "% of sleeve deployed");   calc(r + 6, 2, "=IF(B7>0,B15/B7,0)", "0.0%")
    # B11 risk/share, B13 max risk, B14 shares, B15 capital

    # ── 25-50-25 pyramid ──
    r = 18
    section(ws, r, "25-50-25 Pyramid Split (of max shares)", 4)
    label(r + 1, 1, "T1 pilot (25%)");         calc(r + 1, 2, "=FLOOR(B14*0.25,1)", "#,##0")
    label(r + 2, 1, "T2 core (50%)");          calc(r + 2, 2, "=FLOOR(B14*0.5,1)", "#,##0")
    label(r + 3, 1, "T3 final (25%)");         calc(r + 3, 2, "=B14-B19-B20", "#,##0")

    # ── multi-target R:R ──
    r = 23
    section(ws, r, "Targets & Weighted R:R (edit prices + % to sell)", 6)
    hdr = ["Target", "Price", "% of position", "R:R", "Reward %", "P&L ₹ (at max shares)"]
    rr = r + 1
    for ci, h in enumerate(hdr, start=1):
        cell = ws.cell(row=rr, column=ci, value=h)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
    for i, (tp, pct) in enumerate([(110.0, 33), (120.0, 33), (135.0, 34)], start=1):
        row_i = rr + i
        label(row_i, 1, f"T{i}")
        inp(row_i, 2, tp)
        inp(row_i, 3, pct, "0")
        calc(row_i, 4, f"=IF($B$5-$B$6>0,(B{row_i}-$B$5)/($B$5-$B$6),0)", "0.00")
        calc(row_i, 5, f"=(B{row_i}-$B$5)/$B$5", "0.0%")
        calc(row_i, 6, f"=(B{row_i}-$B$5)*FLOOR($B$14*C{row_i}/100,1)", "#,##0")
    s = rr + 4
    label(s, 1, "Weighted avg R:R", bold=True)
    calc(s, 4, f"=IF(SUM(C{rr+1}:C{rr+3})>0,SUMPRODUCT(D{rr+1}:D{rr+3},C{rr+1}:C{rr+3})/SUM(C{rr+1}:C{rr+3}),0)", "0.00")
    label(s + 1, 1, "Avg exit price", bold=True)
    calc(s + 1, 2, f"=IF(SUM(C{rr+1}:C{rr+3})>0,SUMPRODUCT(B{rr+1}:B{rr+3},C{rr+1}:C{rr+3})/SUM(C{rr+1}:C{rr+3}),0)")
    label(s + 2, 1, "Full-run P&L ₹", bold=True)
    calc(s + 2, 2, f"=SUM(F{rr+1}:F{rr+3})", "#,##0")
    label(s + 3, 1, "Payoff ratio (win/loss)", bold=True)
    calc(s + 3, 2, f"=IF(B13>0,B{s+2}/B13,0)", "0.00")
    label(s + 4, 1, "Break-even win rate", bold=True)
    calc(s + 4, 2, f"=IF(B{s+3}>0,1/(1+B{s+3}),0)", "0.0%")
    autosize(ws, [26, 14, 14, 10, 10, 18])


def build_journal_sheet(ws, cfg):
    """Full trade journal with formula-computed P&L and R-multiples."""
    ws.sheet_view.showGridLines = False
    row = section(ws, 1, "SWING TRADE JOURNAL — add new rows below; P&L / R computed by formula", 12)
    cols = ["Ticker", "Cur", "Result", "Date In", "Date Out", "Qty",
            "Avg Entry", "Avg Exit", "Stop", "P&L", "P&L %", "R-multiple"]
    row = header_row(ws, row, cols)
    for c in cfg.get("closed", []):
        s = c.get("summary", {})
        buys = c.get("buys", [])
        sells = c.get("sells", [])
        qty = sum(b.get("qty", 0) for b in buys)
        stop = c.get("stop")
        vals = [c["ticker"], c.get("currency", ""), c.get("result", ""),
                buys[0]["date"] if buys else "", sells[-1]["date"] if sells else "",
                qty, s.get("avg_entry"), s.get("avg_exit"), stop]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN
            if ci in (7, 8, 9):
                cell.number_format = "#,##0.00"
        # formulas: P&L = (exit-entry)*qty ; P&L% ; R = (exit-entry)/(entry-stop)
        c10 = ws.cell(row=row, column=10, value=f"=(H{row}-G{row})*F{row}")
        c10.number_format = "#,##0.00"
        c10.border = THIN
        c11 = ws.cell(row=row, column=11, value=f"=IF(G{row}>0,(H{row}-G{row})/G{row},0)")
        c11.number_format = "+0.0%;-0.0%"
        c11.border = THIN
        c12 = ws.cell(row=row, column=12, value=f"=IF(AND(ISNUMBER(I{row}),G{row}-I{row}<>0),(H{row}-G{row})/(G{row}-I{row}),\"\")")
        c12.number_format = "+0.00;-0.00"
        c12.border = THIN
        row += 1
    # 10 blank formula-ready rows for future trades
    for _ in range(10):
        for ci in range(1, 10):
            ws.cell(row=row, column=ci).border = THIN
        ws.cell(row=row, column=10, value=f"=IF(F{row}=\"\",\"\",(H{row}-G{row})*F{row})").border = THIN
        ws.cell(row=row, column=11, value=f"=IF(G{row}=\"\",\"\",(H{row}-G{row})/G{row})").border = THIN
        ws.cell(row=row, column=12, value=f"=IF(OR(I{row}=\"\",G{row}-I{row}=0),\"\",(H{row}-G{row})/(G{row}-I{row}))").border = THIN
        row += 1
    autosize(ws, [13, 5, 9, 11, 11, 8, 11, 11, 10, 12, 9, 10])
    ws.freeze_panes = "A3"


def build_dashboard(ws, cfg, fp, buckets, fx, price_note):
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="PORTFOLIO DASHBOARD").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=f"Generated {datetime.now():%Y-%m-%d %H:%M}  |  USD/INR {fx}  |  {price_note}").font = DIM
    row = 4
    row = section(ws, row, "My Portfolio", 4)
    row = header_row(ws, row, ["Bucket", "Invested ₹", "Value ₹", "P&L ₹"])
    tot_i = tot_v = 0.0
    for name, (inv, val) in buckets.items():
        vals = [name, inv or None, val or None, (val - inv) if (inv and val) else None]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = THIN
            if ci > 1:
                cell.number_format = "#,##0"
            if ci == 4 and v is not None:
                pnl_font(cell, v)
        tot_i += inv or 0
        tot_v += val or 0
        row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for ci, v in ((2, tot_i), (3, tot_v), (4, tot_v - tot_i)):
        c = ws.cell(row=row, column=ci, value=v)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        if ci == 4:
            pnl_font(c, v)
    row += 2
    st = fp.get("snapshot_totals", {})
    row = section(ws, row, "Father's Portfolio (see Father tab)", 4)
    for label, v in (("Current value", st.get("current_value")),
                     ("Invested", st.get("invested")),
                     ("P&L", st.get("pnl"))):
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=v)
        c.number_format = "#,##0"
        if label == "P&L" and v is not None:
            pnl_font(c, v)
        row += 1
    autosize(ws, [30, 15, 15, 15])


# ── main ────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(HERE / "Portfolio_Tracker.xlsx"))
    ap.add_argument("--no-fetch", action="store_true", help="skip yfinance")
    args = ap.parse_args()

    cfg = yaml.safe_load((HERE / "positions.yaml").read_text())
    fp_path = HERE / "father_portfolio.yaml"
    fp = yaml.safe_load(fp_path.read_text()) if fp_path.exists() else {}
    fx = cfg.get("networth", {}).get("usd_inr_rate", 95.0)

    prices = Prices(fetch=not args.no_fetch)
    if prices.fetch:
        price_note = "prices: live via yfinance where available"
    elif prices.overrides:
        price_note = f"prices: override file ({getattr(prices, 'override_time', '?')}), yfinance off"
    else:
        price_note = "prices: cached from last YAML sync"

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    z = cfg.get("zerodha_holdings", {})
    im = cfg.get("indmoney_indian", {})
    im_broker = im.get("broker", "INDmoney")
    groups = [
        ("Zerodha — ETFs", z.get("etfs", [])),
        ("Zerodha — REITs / InvITs", z.get("reits_invits", [])),
        (f"{im_broker} — ETFs", im.get("etfs", [])),
    ]
    ws = wb.create_sheet("ETFs & REITs")
    etf_v, etf_i = build_holdings_sheet(ws, "ETFs & REITs", groups, prices)

    ws = wb.create_sheet("Stocks India")
    stk_v, stk_i = build_holdings_sheet(ws, "Stocks", [(f"{im_broker} — Stocks", im.get("stocks", []))], prices)

    ws = wb.create_sheet("Mutual Funds")
    build_mf_sheet(ws, cfg.get("mutual_funds", []), cfg.get("mutual_funds_totals", {}))

    ws = wb.create_sheet("US Holdings")
    us_dormant = cfg.get("us_holdings_dormant", {})
    us_accounts = [("Alpaca (active)", cfg.get("us_holdings", {}))]
    if us_dormant:
        us_accounts.append((f"{us_dormant.get('broker', 'Dormant')} (set-and-forget)", us_dormant))
    us_v, us_i = build_us_sheet(ws, us_accounts, fx, prices)

    ws = wb.create_sheet("Swing & Watchlist")
    build_swing_sheet(ws, cfg, prices)

    ws = wb.create_sheet("Journal")
    build_journal_sheet(ws, cfg)

    ws = wb.create_sheet("Calc — Sizing & RR")
    build_calc_sheet(ws, cfg.get("swing_account", {}).get("sleeve_target_inr", 263453))

    ws = wb.create_sheet("Father")
    build_father_sheet(ws, fp, prices)

    mf_tot = cfg.get("mutual_funds_totals", {})
    buckets = {
        "ETFs & REITs (IN)": (etf_i, etf_v),
        "Stocks (IN)": (stk_i, stk_v),
        "Mutual Funds": (mf_tot.get("total_invested", 0), mf_tot.get("total_current", 0)),
        "US Holdings": (us_i, us_v),
        "Zerodha cash": (cfg.get("networth", {}).get("zerodha_cash", 0),
                         cfg.get("networth", {}).get("zerodha_cash", 0)),
    }
    build_dashboard(ws_dash, cfg, fp, buckets, fx, price_note)

    wb.save(args.out)
    print(f"Wrote {args.out}  ({price_note})")


if __name__ == "__main__":
    main()
